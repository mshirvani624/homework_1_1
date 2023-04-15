import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook , Workbook
import openpyxl

ws=load_workbook("Files/ln.xlsx")["Sheet"]
# quit()
max_row=ws.max_row
data=[]
for i in range(1,max_row+1):
   data.append([ws["A{}".format(i)].value , ws["B{}".format(i)].value])
data=np.array(data)
x=data[:,0]
y=data[:,1]
dy=y[1:]-y[:-1]
dx=x[1:]-x[:-1]
ydot=y.copy()
ydot[1:]=dy/dx
ydot[0]=ydot[1]

dydot=ydot[1:] - ydot[:-1]
# y2dot=dy.copy()
y2dot=y.copy()
y2dot[1:]=dydot/dx

y2dot[0]=y2dot[2]
y2dot[1]=y2dot[2]

plt.plot(x,y,label="ln(x)")
plt.plot(x,ydot,label="1/x")
plt.plot(x,y2dot,label="-1/x^2")
plt.xlabel("X")
plt.ylabel("Y")
plt.title("ln(x) & 1/x & -1/x^2 ")
plt.legend()
plt.show()
wb_save=Workbook()
ws_save=wb_save.active
i=0
for x_i in x :
   ws_save[f"A{i+1}"]=x_i
   ws_save[f"B{i+1}"]=ydot[i]
   ws_save[f"C{i+1}"]=y2dot[i]
   i +=1
wb_save.save("files/ydot_y2dot.xlsx")







